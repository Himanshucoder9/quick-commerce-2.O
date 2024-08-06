import os
import csv
import io
import openpyxl
from django.utils.text import slugify
import requests
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import IntegrityError
from django.http import HttpResponse
from rest_framework.exceptions import NotFound
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.views import APIView
from rest_framework.viewsets import ReadOnlyModelViewSet, ModelViewSet
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from Auth.models import WareHouse
from Warehouse.serializers import (
    TaxSerializer,
    UnitSerializer,
    PackagingTypeSerializer,
    CategorySerializer,
    SimpleSubCategorySerializer,
    SubCategorySerializer,
    SimpleProductSerializer,
    ProductSerializer,
    DetailProductSerializer,
    FullCategorySerializer,
    FullSubCategorySerializer,
    FullProductSerializer,
    ProductDisableSerializer
)
from Warehouse.models import (
    Tax,
    Unit,
    PackagingType,
    Category,
    SubCategory,
    Product
)


class BaseListView(ListAPIView):
    """Base view to handle listing with common functionality."""

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({"message": f"No {self.serializer_class.Meta.model.__name__.lower()} found."},
                            status=status.HTTP_404_NOT_FOUND)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TaxListView(BaseListView):
    serializer_class = TaxSerializer

    def get_queryset(self):
        return Tax.objects.all()


class UnitListView(BaseListView):
    serializer_class = UnitSerializer

    def get_queryset(self):
        return Unit.objects.all()


class PackagingTypeListView(BaseListView):
    serializer_class = PackagingTypeSerializer

    def get_queryset(self):
        return PackagingType.objects.all()


class CategoryBaseListView(BaseListView):
    """Base view for categories to filter by warehouse ID."""

    def get_queryset(self):
        warehouse_id = self.kwargs.get('warehouse_id')
        return Category.objects.filter(warehouse_id=warehouse_id, is_deleted=False)


class SimpleCategoryListView(CategoryBaseListView):
    serializer_class = CategorySerializer


class CategoryListView(CategoryBaseListView):
    serializer_class = CategorySerializer

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        for category in response.data:
            category['subcategories'] = SimpleSubCategorySerializer(
                SubCategory.objects.filter(category_id=category['id'], is_deleted=False),
                many=True, context={'request': request}
            ).data
        return response


class SubCategoryBaseListView(BaseListView):
    """Base view for subcategories to filter by warehouse ID."""

    def get_queryset(self):
        warehouse_id = self.kwargs.get('warehouse_id')
        return SubCategory.objects.filter(warehouse_id=warehouse_id, is_deleted=False)


class SimpleSubCategoryListView(SubCategoryBaseListView):
    serializer_class = SimpleSubCategorySerializer


class SubCategoryListView(SubCategoryBaseListView):
    serializer_class = SubCategorySerializer


class ProductBaseListView(ListAPIView):
    """Base view for products to filter by warehouse ID."""

    def get_queryset(self):
        warehouse_id = self.kwargs.get('warehouse_id')
        return Product.objects.filter(warehouse_id=warehouse_id, is_deleted=False)


class SimpleAllProductListView(ProductBaseListView):
    serializer_class = SimpleProductSerializer


class AllProductListView(ProductBaseListView):
    serializer_class = ProductSerializer


class ProductDetailView(RetrieveAPIView):
    serializer_class = DetailProductSerializer

    def get_queryset(self):
        slug = self.kwargs.get('slug')
        sku_no = self.kwargs.get('sku_no')
        queryset = Product.objects.filter(slug=slug, sku_no=sku_no, is_deleted=False)

        if not queryset.exists():
            raise NotFound("Product not found.")

        return queryset


# Admin ViewSets
class BaseModelViewSet(ModelViewSet):
    """Base viewset to handle common functionality."""

    permission_classes = (IsAuthenticated,)
    http_method_names = ["get", "post", "patch", "delete"]

    def get_object(self):
        slug = self.kwargs.get('slug')
        try:
            return self.queryset.get(warehouse=self.request.user, slug=slug, is_deleted=False)
        except self.queryset.model.DoesNotExist:
            raise NotFound(f"{self.queryset.model.__name__} not found.")

    def destroy(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_deleted = True
        self.object.save()
        return Response(f"{self.queryset.model.__name__} deleted successfully", status=status.HTTP_204_NO_CONTENT)


class CategoryViewSet(BaseModelViewSet):
    queryset = Category.objects.all()
    serializer_class = FullCategorySerializer
    lookup_field = "slug"


class SubCategoryViewSet(BaseModelViewSet):
    queryset = SubCategory.objects.all()
    serializer_class = FullSubCategorySerializer
    lookup_field = "slug"


class ProductViewSet(BaseModelViewSet):
    queryset = Product.objects.all()
    serializer_class = FullProductSerializer
    lookup_field = "slug"

    def get_object(self):
        slug = self.kwargs.get('slug')
        sku_no = self.kwargs.get('sku_no')
        try:
            return Product.objects.get(warehouse=self.request.user, slug=slug, sku_no=sku_no, is_deleted=False)
        except Product.DoesNotExist:
            raise NotFound("Product not found.")

    def update(self, request, *args, **kwargs):
        self.object = self.get_object()
        serializer = self.get_serializer(self.object, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response({"message": "Product updated successfully.", "data": serializer.data})


class ProductDisableView(APIView):
    permission_classes = [IsAuthenticated]

    def get_product(self, slug, sku_no, user):
        try:
            return Product.objects.get(slug=slug, sku_no=sku_no, warehouse=user)
        except Product.DoesNotExist:
            return None

    def put(self, request, slug, sku_no):
        product = self.get_product(slug, sku_no, request.user)
        if product:
            product.is_active = False
            product.save()
            serializer = ProductDisableSerializer(product)
            return Response({"message": "Product disabled successfully.", "product": serializer.data})
        return Response({"error": "Product not found."}, status=status.HTTP_404_NOT_FOUND)


# Bulk
class CategoryBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Check if a file is provided
        if 'file' not in request.FILES:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        data = []

        # Define expected column names
        expected_columns = ['title', 'image_url']

        # Handle CSV files
        if file.name.endswith('.csv'):
            decoded_file = file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            # Validate columns
            if reader.fieldnames != expected_columns:
                return Response(
                    {"error": f"Invalid CSV columns. Expected: {expected_columns}, Got: {reader.fieldnames}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for row in reader:
                for field in expected_columns:
                    if field not in row or not row[field]:
                        return Response(
                            {"error": f"Missing or empty required field '{field}' in CSV."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                data.append(row)

        # Handle Excel files
        elif file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Validate column names in the header row
            header = [cell.value for cell in sheet[1]]  # Read header row
            if header != expected_columns:
                return Response(
                    {"error": f"Invalid Excel columns. Expected: {expected_columns}, Got: {header}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(header, row))
                for field in expected_columns:
                    if field not in row_data or not row_data[field]:
                        return Response(
                            {"error": f"Missing or empty required field '{field}' in Excel."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                data.append(row_data)
        else:
            return Response({"error": "Unsupported file format. Only .csv and .xlsx are allowed."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Get the warehouse instance for the authenticated user
        try:
            warehouse = WareHouse.objects.get(id=request.user.id)
        except WareHouse.DoesNotExist:
            return Response({"error": "Authenticated user does not have an associated warehouse."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Prepare to bulk create or update categories
        categories_to_create = []
        categories_to_update = []
        response_data = []

        for item in data:
            slug = slugify(item['title'])
            image_name = f"{slug}.webp"  # Use appropriate image extension
            image_path = f"category/image/{image_name}"

            # Download the image from the provided URL
            try:
                response = requests.get(item['image_url'])
                response.raise_for_status()  # Raise an error for bad responses

                # Save the image to the default storage
                image_content = ContentFile(response.content)
                image_file_name = default_storage.save(image_path, image_content)  # Save image to media directory
            except Exception as e:
                return Response({"error": f"Failed to download image from {item['image_url']}: {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST)

            # Prepare category data for bulk create or update
            category_data = {
                'title': item['title'],
                'warehouse': warehouse,
                'image': image_file_name,
                'slug': slug,
                'is_deleted': False
            }

            # Check if the category already exists
            try:
                category = Category.objects.get(title=item['title'], warehouse=warehouse)

                # If the category exists, prepare it for update
                category_data['id'] = category.id
                categories_to_update.append(category_data)
            except Category.DoesNotExist:
                # If it doesn't exist, prepare it for creation
                categories_to_create.append(category_data)

        # Perform bulk create
        if categories_to_create:
            try:
                created_categories = Category.objects.bulk_create(
                    [Category(**data) for data in categories_to_create]
                )
                for category in created_categories:
                    response_data.append({
                        'id': category.id,
                        'warehouse': category.warehouse.id,
                        'title': category.title,
                        'image': category.image.url,
                        'slug': category.slug,
                        'created_at': category.created_at,
                        'updated_at': category.updated_at
                    })
            except Exception as e:
                return Response({"error": f"Error during bulk creation: {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST)

        # Perform bulk update
        if categories_to_update:
            try:
                for category_data in categories_to_update:
                    category = Category.objects.get(id=category_data['id'])

                    # Delete the old image if it exists
                    if category.image:
                        old_image_path = category.image.path
                        if os.path.exists(old_image_path):
                            os.remove(old_image_path)  # Delete the old image file

                    # Update category fields
                    category.image = category_data['image']
                    category.slug = category_data['slug']
                    category.is_deleted = category_data['is_deleted']
                    category.save()  # Save the updated category

                    # Add to response data
                    response_data.append({
                        'id': category.id,
                        'warehouse': category.warehouse.id,
                        'title': category.title,
                        'image': category.image.url,
                        'slug': category.slug,
                        'created_at': category.created_at,
                        'updated_at': category.updated_at
                    })
            except Exception as e:
                return Response({"error": f"Error during bulk update: {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST)

        # Return a JSON response with the processed data
        return Response({"message": "Categories processed successfully.", "data": response_data},
                        status=status.HTTP_200_OK)




class SubCategoryBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Check if a file is provided
        if 'file' not in request.FILES:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        data = []

        # Define expected column names
        expected_columns = ['title', 'category_id', 'image_url']

        # Handle CSV files
        if file.name.endswith('.csv'):
            decoded_file = file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            # Validate columns
            if reader.fieldnames != expected_columns:
                return Response(
                    {"error": f"Invalid CSV columns. Expected: {expected_columns}, Got: {reader.fieldnames}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for row in reader:
                for field in expected_columns:
                    if field not in row or not row[field]:
                        return Response(
                            {"error": f"Missing or empty required field '{field}' in CSV."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                data.append(row)

        # Handle Excel files
        elif file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Validate column names in the header row
            header = [cell.value for cell in sheet[1]]  # Read header row
            if header != expected_columns:
                return Response(
                    {"error": f"Invalid Excel columns. Expected: {expected_columns}, Got: {header}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(header, row))
                for field in expected_columns:
                    if field not in row_data or not row_data[field]:
                        return Response(
                            {"error": f"Missing or empty required field '{field}' in Excel."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                data.append(row_data)
        else:
            return Response({"error": "Unsupported file format. Only .csv and .xlsx are allowed."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Get the warehouse instance for the authenticated user
        try:
            warehouse = WareHouse.objects.get(id=request.user.id)
        except WareHouse.DoesNotExist:
            return Response({"error": "Authenticated user does not have an associated warehouse."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Prepare to bulk create or update subcategories
        subcategories_to_create = []
        subcategories_to_update = []
        response_data = []

        for item in data:
            slug = slugify(item['title'])
            image_name = f"{slug}.webp"  # Use appropriate image extension
            image_path = f"subcategory/image/{image_name}"

            # Download the image from the provided URL
            try:
                response = requests.get(item['image_url'])
                response.raise_for_status()  # Raise an error for bad responses

                # Save the image to the default storage
                image_content = ContentFile(response.content)
                image_file_name = default_storage.save(image_path, image_content)  # Save image to media directory
            except Exception as e:
                return Response({"error": f"Failed to download image from {item['image_url']}: {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST)

            # Prepare subcategory data
            subcategory_data = {
                'title': item['title'],
                'warehouse': warehouse,
                'image': image_file_name,
                'slug': slug,
                'is_deleted': False,
                'category': None  # Placeholder for category, will set below
            }

            # Check if the category exists
            try:
                category = Category.objects.get(id=item['category_id'], warehouse=warehouse)
                subcategory_data['category'] = category  # Set the category for creation

                # Check if the subcategory already exists
                try:
                    subcategory = SubCategory.objects.get(title=item['title'], warehouse=warehouse, category=category)

                    # If the subcategory exists, prepare it for update
                    subcategory.image = image_file_name
                    subcategory.slug = slug
                    subcategory.is_deleted = False  # Reset is_deleted to False
                    subcategory.save()
                    response_data.append({
                        'id': subcategory.id,
                        'warehouse': subcategory.warehouse.id,
                        'category': subcategory.category.id,
                        'title': subcategory.title,
                        'image': subcategory.image.url,
                        'slug': subcategory.slug,
                        'created_at': subcategory.created_at,
                        'updated_at': subcategory.updated_at
                    })

                except SubCategory.DoesNotExist:
                    # If it doesn't exist, prepare it for creation
                    subcategories_to_create.append(subcategory_data)

            except Category.DoesNotExist:
                return Response(
                    {"error": f"Category with id '{item['category_id']}' does not exist."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Perform bulk create for new subcategories
        if subcategories_to_create:
            try:
                created_subcategories = SubCategory.objects.bulk_create(
                    [SubCategory(**data) for data in subcategories_to_create]
                )
                for subcategory in created_subcategories:
                    response_data.append({
                        'id': subcategory.id,
                        'warehouse': subcategory.warehouse.id,
                        'category': subcategory.category.id,
                        'title': subcategory.title,
                        'image': subcategory.image.url,
                        'slug': subcategory.slug,
                        'created_at': subcategory.created_at,
                        'updated_at': subcategory.updated_at
                    })
            except Exception as e:
                return Response({"error": f"Error during bulk creation: {str(e)}"},
                                status=status.HTTP_400_BAD_REQUEST)

        # Return a JSON response with the processed data
        return Response({"message": "Subcategories processed successfully.", "data": response_data},
                        status=status.HTTP_200_OK)



